# Mengimpor library yang diperlukan
import pandas as pd
import matplotlib.pyplot as plt
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def compare_algorithms(apriori_file, fpgrowth_file, output_file):
    """
    Fungsi untuk membandingkan hasil algoritma Apriori dan FP-Growth
    
    Parameters:
        apriori_file (str): Path file Excel hasil Apriori
        fpgrowth_file (str): Path file Excel hasil FP-Growth
        output_file (str): Path file output untuk menyimpan hasil perbandingan
    """
    
    # Membaca hasil dari kedua metode menggunakan pandas
    try:
        apriori_results = pd.read_excel(apriori_file, sheet_name='Metode Apriori')
        fpgrowth_results = pd.read_excel(fpgrowth_file, sheet_name='Metode FP-Growth')
    except Exception as e:
        print(f"Error reading input files: {e}")
        return

    # Fungsi internal untuk mengekstrak metrik dari dataframe hasil
    def extract_metrics(df):
        """
        Mengekstrak metrik dari dataframe hasil algoritma
        
        Parameters:
            df (DataFrame): DataFrame hasil algoritma
            
        Returns:
            Series: Seri pandas berisi metrik yang diekstrak
        """
        try:
            # Mencari baris yang berisi kata 'Metrics'
            metrics_mask = df.iloc[:, 0].astype(str).str.contains('Metrics', na=False)
            if not metrics_mask.any():
                return pd.Series(dtype=float)
                
            metrics_start = metrics_mask.idxmax()
            # Mengambil data metrik (7 baris setelah 'Metrics')
            metrics_data = df.iloc[metrics_start+1:metrics_start+8, :2].dropna()
            
            # Validasi data metrics
            if metrics_data.empty or metrics_data.shape[1] < 2:
                return pd.Series(dtype=float)
                
            # Membersihkan nama kolom
            metrics_data.columns = ['Metric', 'Value']
            metrics_data['Metric'] = metrics_data['Metric'].astype(str).str.strip()
            
            # Konversi nilai ke float jika memungkinkan
            metrics_data['Value'] = pd.to_numeric(metrics_data['Value'], errors='coerce')
            
            return metrics_data.set_index('Metric')['Value']
        except Exception as e:
            print(f"Error extracting metrics: {e}")
            return pd.Series(dtype=float)

    # Mengekstrak metrik untuk kedua algoritma
    apriori_metrics = extract_metrics(apriori_results)
    fpgrowth_metrics = extract_metrics(fpgrowth_results)
    
    # Fungsi untuk mendapatkan nilai dengan handling NaN
    def get_metric(metrics, key, default=0):
        value = metrics.get(key, default)
        return default if pd.isna(value) else value
    
    # Membuat dictionary untuk data perbandingan
    comparison_data = {
        'Metrik': [
            'Waktu Eksekusi (detik)', 
            'Jumlah Rule Ditemukan',
            'Rata-rata Akurasi',
            'Rata-rata Lift',
            'Min Support',
            'Min Confidence'
        ],
        'Apriori': [
            get_metric(apriori_metrics, 'Waktu Eksekusi (detik)'),
            get_metric(apriori_metrics, 'Jumlah Rule Ditemukan'),
            get_metric(apriori_metrics, 'Rata-rata Akurasi'),
            get_metric(apriori_metrics, 'Rata-rata Lift'),
            get_metric(apriori_metrics, 'Min Support'),
            get_metric(apriori_metrics, 'Min Confidence')
        ],
        'FP-Growth': [
            get_metric(fpgrowth_metrics, 'Waktu Eksekusi (detik)'),
            get_metric(fpgrowth_metrics, 'Jumlah Rule Ditemukan'),
            get_metric(fpgrowth_metrics, 'Rata-rata Akurasi'),
            get_metric(fpgrowth_metrics, 'Rata-rata Lift'),
            get_metric(fpgrowth_metrics, 'Min Support'),
            get_metric(fpgrowth_metrics, 'Min Confidence')
        ]
    }
    
    # Membuat DataFrame dari data perbandingan
    comparison_df = pd.DataFrame(comparison_data)
    
    # Menghitung persentase perbedaan antara FP-Growth dan Apriori
    comparison_df['Perbedaan'] = comparison_df.apply(
        lambda x: f"{((x['FP-Growth'] - x['Apriori'])/x['Apriori'])*100:.1f}%" 
        if isinstance(x['Apriori'], (int, float)) and isinstance(x['FP-Growth'], (int, float)) and x['Apriori'] != 0 
        else 'N/A', axis=1)
    
    # Membuat figure untuk visualisasi dengan ukuran 15x15 inci
    plt.figure(figsize=(15, 15))
    
    # 1. Grafik Bar untuk Waktu Eksekusi
    plt.subplot(3, 2, 1)
    apriori_time = get_metric(apriori_metrics, 'Waktu Eksekusi (detik)', 0)
    fpgrowth_time = get_metric(fpgrowth_metrics, 'Waktu Eksekusi (detik)', 0)
    plt.bar(['Apriori', 'FP-Growth'], [apriori_time, fpgrowth_time],
            color=['#1f77b4', '#ff7f0e'])
    plt.title('Perbandingan Waktu Eksekusi', fontsize=12, fontweight='bold')
    plt.ylabel('Detik', fontsize=10)
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    
    # 2. Grafik Bar untuk Jumlah Rule
    plt.subplot(3, 2, 2)
    apriori_rules = get_metric(apriori_metrics, 'Jumlah Rule Ditemukan', 0)
    fpgrowth_rules = get_metric(fpgrowth_metrics, 'Jumlah Rule Ditemukan', 0)
    plt.bar(['Apriori', 'FP-Growth'], [apriori_rules, fpgrowth_rules],
            color=['#1f77b4', '#ff7f0e'])
    plt.title('Perbandingan Jumlah Rule', fontsize=12, fontweight='bold')
    plt.ylabel('Jumlah', fontsize=10)
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    
    # 3. Grafik Pie untuk Akurasi (hanya jika data valid)
    plt.subplot(3, 2, 3)
    apriori_acc = get_metric(apriori_metrics, 'Rata-rata Akurasi', 0)
    fpgrowth_acc = get_metric(fpgrowth_metrics, 'Rata-rata Akurasi', 0)
    
    if apriori_acc > 0 or fpgrowth_acc > 0:  # Hanya plot jika ada data valid
        plt.pie([apriori_acc, fpgrowth_acc],
                labels=['Apriori', 'FP-Growth'],
                colors=['#1f77b4', '#ff7f0e'],
                autopct='%1.1f%%',
                startangle=90,
                textprops={'fontsize': 10})
        plt.title('Distribusi Rata-rata Akurasi', fontsize=12, fontweight='bold')
    else:
        plt.text(0.5, 0.5, 'Data Akurasi Tidak Tersedia', 
                ha='center', va='center', fontsize=12)
        plt.title('Distribusi Rata-rata Akurasi', fontsize=12, fontweight='bold')
    
    # 4. Grafik Line untuk Lift
    plt.subplot(3, 2, 4)
    apriori_lift = get_metric(apriori_metrics, 'Rata-rata Lift', 0)
    fpgrowth_lift = get_metric(fpgrowth_metrics, 'Rata-rata Lift', 0)
    plt.plot(['Apriori', 'FP-Growth'], [apriori_lift, fpgrowth_lift],
             marker='o',
             color='#2ca02c',
             linewidth=2)
    plt.title('Perbandingan Rata-rata Lift', fontsize=12, fontweight='bold')
    plt.ylabel('Lift', fontsize=10)
    plt.grid(linestyle='--', alpha=0.7)
    
    # 5. Grafik Bar Horizontal untuk Support dan Confidence
    plt.subplot(3, 2, 5)
    metrics = ['Min Support', 'Min Confidence']
    apriori_vals = [get_metric(apriori_metrics, m, 0) for m in metrics]
    fpgrowth_vals = [get_metric(fpgrowth_metrics, m, 0) for m in metrics]
    
    y = range(len(metrics))
    plt.barh([y_pos - 0.2 for y_pos in y], apriori_vals, height=0.4, label='Apriori', color='#1f77b4')
    plt.barh([y_pos + 0.2 for y_pos in y], fpgrowth_vals, height=0.4, label='FP-Growth', color='#ff7f0e')
    
    plt.yticks(y, metrics)
    plt.title('Parameter Algoritma', fontsize=12, fontweight='bold')
    plt.xlabel('Nilai', fontsize=10)
    plt.legend(fontsize=9)
    plt.grid(axis='x', linestyle='--', alpha=0.7)
    
    # 6. Grafik Gabungan untuk Waktu vs Jumlah Rule
    plt.subplot(3, 2, 6)
    ax1 = plt.gca()
    ax2 = ax1.twinx()
    
    ax1.plot(['Apriori', 'FP-Growth'], [apriori_time, fpgrowth_time],
             color='#1f77b4',
             marker='o',
             label='Waktu (detik)')
    ax2.plot(['Apriori', 'FP-Growth'], [apriori_rules, fpgrowth_rules],
             color='#ff7f0e',
             marker='s',
             label='Jumlah Rule')
    
    ax1.set_ylabel('Waktu (detik)', color='#1f77b4')
    ax1.tick_params(axis='y', colors='#1f77b4')
    ax2.set_ylabel('Jumlah Rule', color='#ff7f0e')
    ax2.tick_params(axis='y', colors='#ff7f0e')
    plt.title('Waktu vs Jumlah Rule', fontsize=12, fontweight='bold')
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left', fontsize=8)
    plt.grid(axis='x', linestyle='--', alpha=0.7)
    
    # Menyesuaikan layout agar tidak tumpang tindih
    plt.tight_layout()
    
    # Membuat direktori output jika belum ada
    os.makedirs(output_dir, exist_ok=True)
    
    # Menyimpan gambar visualisasi
    image_path = os.path.join(output_dir, 'hasil_perbandingan.jpg')
    plt.savefig(image_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    # Membuat Excel writer untuk menyimpan hasil
    try:
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        comparison_df.to_excel(writer, sheet_name='Perbandingan', index=False)
        
        # Formatting Excel
        workbook = writer.book
        worksheet = writer.sheets['Perbandingan']
        
        # Style untuk header
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))
        
        # Menerapkan format ke header
        for col in range(1, 5):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Menerapkan format ke isi tabel
        for row in range(2, len(comparison_df)+2):
            for col in range(1, 5):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                if col > 1:
                    cell.alignment = Alignment(horizontal='right')
        
        # Menyesuaikan lebar kolom otomatis
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
        
        # Menambahkan gambar visualisasi ke Excel
        try:
            img = Image(image_path)
            img.height = 1200 * 0.75
            img.width = 1000 * 0.75
            img.anchor = 'A' + str(len(comparison_df) + 4)
            worksheet.add_image(img)
        except Exception as e:
            print(f"Error adding image to Excel: {e}")
        
        # Memperbaiki header tabel
        worksheet['A1'].value = 'METRIK'
        worksheet['B1'].value = 'APRIORI'
        worksheet['C1'].value = 'FP-GROWTH'
        worksheet['D1'].value = 'PERBEDAAN'
        
        # Menambahkan teks analisis
        worksheet['A' + str(len(comparison_df)+3)] = 'ANALISIS:'
        worksheet['A' + str(len(comparison_df)+3)].font = Font(bold=True)
        
        analysis_text = [
            "1. FP-Growth umumnya lebih cepat daripada Apriori untuk dataset besar",
            "2. Jumlah rule yang dihasilkan tergantung pada min support dan confidence",
            "3. Lift > 1 menunjukkan hubungan positif antar item",
            "4. Akurasi tinggi menunjukkan rule yang dihasilkan lebih reliable"
        ]
        
        for i, text in enumerate(analysis_text, start=1):
            worksheet['K' + str(len(comparison_df)+3+i)] = text
        
        # Menyimpan file Excel
        writer.close()
        
        # Mencetak informasi output
        print(f"Hasil perbandingan disimpan di: {output_file}")
        print(f"Gambar visualisasi disimpan di: {image_path}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")

if __name__ == "__main__":
    # Mendefinisikan path file input dan output
    apriori_result_file = 'data_temp/tahap4_hasil_perhitungan.xlsx'
    fpgrowth_result_file = 'data_temp/tahap5_hasil_perhitungan.xlsx'
    output_comparison_file = 'data_temp/tahap6_hasil_perbandingan.xlsx'
    output_dir = 'data_temp/result_output'
    
    # Menjalankan fungsi perbandingan
    compare_algorithms(apriori_result_file, fpgrowth_result_file, output_comparison_file)